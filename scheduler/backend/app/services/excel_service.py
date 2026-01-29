"""
Excel Import/Export Service

Handles parsing of existing Excel schedules and exporting generated schedules.
"""

import openpyxl
import openpyxl.comments
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
import pandas as pd
from datetime import datetime, time, date
from typing import Dict, List, Any, Optional, Tuple
import re
from pathlib import Path
from io import BytesIO

from app.models import (
    Student, Staff, SchoolClass, CareTime, WorkHour, Absence,
    StaffRole, ScheduleType, GradeGroup
)


class ExcelParseError(Exception):
    """Raised when Excel parsing fails."""
    pass


class ExcelImportService:
    """
    Service for importing data from Excel schedules.

    Handles the specific format of "Schema att maila Joel.xlsx".
    """

    def __init__(self):
        self.students_data: Dict[str, Any] = {}
        self.staff_data: Dict[str, Any] = {}
        self.classes_data: Dict[str, Any] = {}

    def parse_schedule_excel(self, filepath: str) -> Dict[str, Any]:
        """
        Parse the Excel schedule file and extract all data.

        Args:
            filepath: Path to Excel file

        Returns:
            Dictionary with parsed students, staff, classes, etc.

        Raises:
            ExcelParseError: If parsing fails
        """
        try:
            # Load all sheets
            excel_file = pd.ExcelFile(filepath, engine='openpyxl')
            sheets = {sheet_name: excel_file.parse(sheet_name) for sheet_name in excel_file.sheet_names}

            # Parse each day's sheet
            weekdays = ['måndag', 'tisdag', 'onsdag', 'torsdag', 'fredag']

            for sheet_name, df in sheets.items():
                # Determine weekday from sheet name
                weekday = self._extract_weekday_from_sheet_name(sheet_name)
                if weekday is not None:
                    self._parse_day_sheet(df, weekday)

            # Build complete data structure
            return {
                'students': self._build_students_list(),
                'staff': self._build_staff_list(),
                'classes': self._build_classes_list(),
            }

        except Exception as e:
            raise ExcelParseError(f"Failed to parse Excel file: {str(e)}")

    def _extract_weekday_from_sheet_name(self, sheet_name: str) -> Optional[int]:
        """
        Extract weekday number (0-6) from sheet name.

        Args:
            sheet_name: Name of Excel sheet (e.g., "mån 12 jan")

        Returns:
            Weekday number (0=Monday, 6=Sunday) or None if not a weekday sheet
        """
        sheet_lower = sheet_name.lower()

        weekday_map = {
            'mån': 0, 'måndag': 0,
            'tis': 1, 'tisdag': 1,
            'ons': 2, 'onsdag': 2,
            'tors': 3, 'torsdag': 3, 'tor': 3,
            'fre': 4, 'fredag': 4,
            'lör': 5, 'lördag': 5,
            'sön': 6, 'söndag': 6,
        }

        for day_name, day_num in weekday_map.items():
            if day_name in sheet_lower:
                return day_num

        return None

    def _parse_day_sheet(self, df: pd.DataFrame, weekday: int) -> None:
        """
        Parse a single day's sheet and extract student/staff data.

        Args:
            df: DataFrame for the day
            weekday: Weekday number (0-6)
        """
        # Find sections: "Stjärnan låg" (grades 1-3) and "Stjärnan mellan" (grades 4-6)

        current_section = None
        current_grade_group = None

        for idx, row in df.iterrows():
            # Skip completely empty rows
            if row.isna().all():
                continue

            # Detect section headers
            row_str = ' '.join([str(val) for val in row.values if pd.notna(val)]).lower()

            if 'stjärnan låg' in row_str or 'stj' in row_str and 'l' in row_str:
                current_grade_group = GradeGroup.GRADES_1_3
                continue
            elif 'stjärnan mellan' in row_str or 'mellan' in row_str:
                current_grade_group = GradeGroup.GRADES_4_6
                continue

            # Skip header rows
            if any(header in row_str for header in ['start', 'slut', 'fritids', 'skola']):
                continue

            # Try to extract student data
            student_name = self._extract_student_name(row)
            if student_name and current_grade_group:
                self._process_student_row(row, student_name, weekday, current_grade_group)

            # Try to extract staff data
            staff_name = self._extract_staff_name(row)
            if staff_name:
                self._process_staff_row(row, staff_name, weekday)

    def _extract_student_name(self, row: pd.Series) -> Optional[str]:
        """
        Extract student name from row.

        Returns name if valid student row, None otherwise.
        """
        # Look for name in first few columns
        for col_idx in range(min(5, len(row))):
            val = row.iloc[col_idx]
            if pd.isna(val):
                continue

            val_str = str(val).strip()

            # Check if it looks like a name (alphabetic, not a time)
            if val_str and val_str[0].isupper() and not ':' in val_str:
                # Exclude common non-name values
                exclude = ['klass', 'personal', 'unnamed', 'nan', 'x']
                if not any(exc in val_str.lower() for exc in exclude):
                    return val_str

        return None

    def _extract_staff_name(self, row: pd.Series) -> Optional[str]:
        """
        Extract staff name from row (usually in later columns).
        """
        # Staff names typically appear in columns after student data
        # Look for pattern: Name followed by times

        for col_idx in range(len(row) - 5, len(row)):
            if col_idx < 0:
                continue

            val = row.iloc[col_idx]
            if pd.isna(val):
                continue

            val_str = str(val).strip()

            # Check if looks like staff name
            if val_str and val_str[0].isupper() and len(val_str) > 2:
                # Check if followed by time values
                has_times = False
                for next_col in range(col_idx + 1, min(col_idx + 4, len(row))):
                    next_val = str(row.iloc[next_col])
                    if ':' in next_val:  # Time format
                        has_times = True
                        break

                if has_times:
                    return val_str

        return None

    def _process_student_row(
        self,
        row: pd.Series,
        student_name: str,
        weekday: int,
        grade_group: GradeGroup
    ) -> None:
        """
        Process a student row and extract care times, preferences, etc.
        """
        # Initialize student data if not exists
        if student_name not in self.students_data:
            # Determine grade from grade_group
            grade = 1 if grade_group == GradeGroup.GRADES_1_3 else 4

            self.students_data[student_name] = {
                'first_name': student_name.split()[0] if ' ' in student_name else student_name,
                'last_name': student_name.split()[-1] if ' ' in student_name else '',
                'grade': grade,
                'grade_group': grade_group,
                'care_times': {},
                'has_special_needs': False,
                'preferred_staff': [],
                'requires_double_staffing': False,
            }

        # Extract care times (Start and Slut columns)
        start_time = self._extract_time_from_row(row, ['start', 'början'])
        end_time = self._extract_time_from_row(row, ['slut', 'slutar'])

        if start_time and end_time:
            self.students_data[student_name]['care_times'][weekday] = {
                'start': start_time,
                'end': end_time,
            }

        # Extract assigned staff (Fritids column)
        fritids_staff = self._extract_staff_from_row(row, ['fritids'])
        if fritids_staff:
            if fritids_staff not in self.students_data[student_name]['preferred_staff']:
                self.students_data[student_name]['preferred_staff'].append(fritids_staff)

        # Check for special markers (star, double staffing)
        row_str = ' '.join([str(val) for val in row.values if pd.notna(val)])
        if '*' in row_str or 'stjärn' in row_str.lower():
            self.students_data[student_name]['has_special_needs'] = True

    def _process_staff_row(self, row: pd.Series, staff_name: str, weekday: int) -> None:
        """
        Process a staff row and extract working hours.
        """
        # Initialize staff data if not exists
        if staff_name not in self.staff_data:
            self.staff_data[staff_name] = {
                'first_name': staff_name.split()[0] if ' ' in staff_name else staff_name,
                'last_name': staff_name.split()[-1] if ' ' in staff_name else '',
                'role': self._infer_role(staff_name),
                'work_hours': {},
                'care_certifications': [],
            }

        # Extract work times
        start_time = self._extract_time_from_row(row, ['start'])
        end_time = self._extract_time_from_row(row, ['slut', 'end'])

        if start_time and end_time:
            self.staff_data[staff_name]['work_hours'][weekday] = {
                'start': start_time,
                'end': end_time,
            }

    def _extract_time_from_row(self, row: pd.Series, keywords: List[str]) -> Optional[str]:
        """
        Extract time value (HH:MM) from row.

        Args:
            row: DataFrame row
            keywords: Keywords to identify the time column

        Returns:
            Time string in HH:MM format or None
        """
        # First, try to find column with keyword in header
        for col_name in row.index:
            col_lower = str(col_name).lower()
            if any(kw in col_lower for kw in keywords):
                val = row[col_name]
                time_str = self._parse_time_value(val)
                if time_str:
                    return time_str

        # If not found, look for time patterns in values
        for val in row.values:
            if pd.isna(val):
                continue
            time_str = self._parse_time_value(val)
            if time_str:
                return time_str

        return None

    def _parse_time_value(self, val: Any) -> Optional[str]:
        """
        Parse a value and extract time in HH:MM format.

        Handles:
        - datetime.time objects
        - Strings like "08:30:00"
        - Strings like "8:30"
        - Excel time serials
        """
        if pd.isna(val):
            return None

        # Handle datetime.time objects
        if isinstance(val, time):
            return val.strftime('%H:%M')

        # Handle string representations
        val_str = str(val).strip()

        # Match HH:MM or HH:MM:SS patterns
        time_match = re.search(r'(\d{1,2}):(\d{2})', val_str)
        if time_match:
            hour = int(time_match.group(1))
            minute = int(time_match.group(2))
            return f"{hour:02d}:{minute:02d}"

        # Handle Excel time serials (0.0 to 1.0 representing fraction of day)
        try:
            val_float = float(val)
            if 0 <= val_float < 1:
                # Convert to time
                total_seconds = int(val_float * 24 * 60 * 60)
                hours = total_seconds // 3600
                minutes = (total_seconds % 3600) // 60
                return f"{hours:02d}:{minutes:02d}"
        except (ValueError, TypeError):
            pass

        return None

    def _extract_staff_from_row(self, row: pd.Series, keywords: List[str]) -> Optional[str]:
        """
        Extract staff name from row based on keywords.
        """
        for col_name in row.index:
            col_lower = str(col_name).lower()
            if any(kw in col_lower for kw in keywords):
                val = row[col_name]
                if pd.notna(val):
                    val_str = str(val).strip()
                    # Exclude 'X' which means no staff
                    if val_str and val_str.upper() != 'X' and len(val_str) > 1:
                        return val_str
        return None

    def _infer_role(self, staff_name: str) -> StaffRole:
        """
        Infer staff role from name or context.

        Default to ASSISTANT since most staff are elevassistenter.
        """
        # Could be enhanced with a lookup table or keyword matching
        # For now, default to assistant
        return StaffRole.ASSISTANT

    def _build_students_list(self) -> List[Dict[str, Any]]:
        """
        Build final list of students from accumulated data.
        """
        students = []

        for idx, (name, data) in enumerate(self.students_data.items(), start=1):
            # Generate personal number (placeholder - would need real data)
            personal_number = f"20{10 + data['grade']}{idx:04d}-0000"

            student = {
                'personal_number': personal_number,
                'first_name': data['first_name'],
                'last_name': data['last_name'],
                'grade': data['grade'],
                'grade_group': data['grade_group'],
                'has_care_needs': data['has_special_needs'],
                'care_requirements': [],  # Would need to extract from notes
                'preferred_staff': data['preferred_staff'],
                'requires_double_staffing': data['requires_double_staffing'],
                'care_times': data['care_times'],
            }

            students.append(student)

        return students

    def _build_staff_list(self) -> List[Dict[str, Any]]:
        """
        Build final list of staff from accumulated data.
        """
        staff_list = []

        for idx, (name, data) in enumerate(self.staff_data.items(), start=1):
            # Generate personal number (placeholder)
            personal_number = f"197{idx % 10}{idx:04d}-0000"

            staff = {
                'personal_number': personal_number,
                'first_name': data['first_name'],
                'last_name': data['last_name'],
                'role': data['role'],
                'care_certifications': data['care_certifications'],
                'work_hours': data['work_hours'],
            }

            staff_list.append(staff)

        return staff_list

    def _build_classes_list(self) -> List[Dict[str, Any]]:
        """
        Build list of school classes.

        For now, create standard classes based on grade groups found.
        """
        classes = [
            {'name': 'Klass 1-3A', 'grade_group': GradeGroup.GRADES_1_3},
            {'name': 'Klass 1-3B', 'grade_group': GradeGroup.GRADES_1_3},
            {'name': 'Klass 1-3C', 'grade_group': GradeGroup.GRADES_1_3},
            {'name': 'Klass 1-3D', 'grade_group': GradeGroup.GRADES_1_3},
            {'name': 'Klass 1-3E', 'grade_group': GradeGroup.GRADES_1_3},
            {'name': 'Klass 4-6A', 'grade_group': GradeGroup.GRADES_4_6},
            {'name': 'Klass 4-6B', 'grade_group': GradeGroup.GRADES_4_6},
            {'name': 'Klass 4-6C', 'grade_group': GradeGroup.GRADES_4_6},
            {'name': 'Klass 4-6D', 'grade_group': GradeGroup.GRADES_4_6},
        ]

        return classes


class ExcelExportService:
    """
    Service for exporting generated schedules to Excel format.
    """

    def export_schedule_to_excel(
        self,
        schedule: 'Schedule',
        filepath: str,
        include_staff_summary: bool = True
    ) -> None:
        """
        Export a generated schedule to Excel file.

        Args:
            schedule: Schedule object to export
            filepath: Output file path
            include_staff_summary: Whether to include staff summary sheet
        """
        # Create workbook
        wb = openpyxl.Workbook()
        wb.remove(wb.active)  # Remove default sheet

        # Create sheets for each weekday
        weekdays = ['Måndag', 'Tisdag', 'Onsdag', 'Torsdag', 'Fredag']

        for weekday_num, weekday_name in enumerate(weekdays):
            ws = wb.create_sheet(title=weekday_name)
            self._populate_day_sheet(ws, schedule, weekday_num)

        # Add staff summary sheet if requested
        if include_staff_summary:
            ws_summary = wb.create_sheet(title='Personal Sammanfattning')
            self._populate_staff_summary(ws_summary, schedule)

        # Save workbook
        wb.save(filepath)

    def _populate_day_sheet(self, ws, schedule: 'Schedule', weekday: int) -> None:
        """
        Populate a single day's sheet with assignments.
        """
        # Add header
        ws['A1'] = 'Personal'
        ws['B1'] = 'Elev'
        ws['C1'] = 'Starttid'
        ws['D1'] = 'Sluttid'
        ws['E1'] = 'Typ'

        # Get assignments for this weekday
        day_assignments = [a for a in schedule.assignments if a.weekday == weekday]

        # Write assignments
        row = 2
        for assignment in day_assignments:
            ws[f'A{row}'] = assignment.staff.full_name
            ws[f'B{row}'] = assignment.student.full_name if assignment.student else 'Klass'
            ws[f'C{row}'] = assignment.start_time
            ws[f'D{row}'] = assignment.end_time
            ws[f'E{row}'] = assignment.assignment_type.value
            row += 1

    def _populate_staff_summary(self, ws, schedule: 'Schedule') -> None:
        """
        Create summary sheet showing total hours per staff member.
        """
        ws['A1'] = 'Personal'
        ws['B1'] = 'Totala timmar'
        ws['C1'] = 'Antal tilldelningar'

        # Calculate totals per staff
        staff_totals = {}

        for assignment in schedule.assignments:
            staff_name = assignment.staff.full_name

            if staff_name not in staff_totals:
                staff_totals[staff_name] = {'hours': 0, 'count': 0}

            # Calculate hours for this assignment
            start = datetime.strptime(assignment.start_time, '%H:%M')
            end = datetime.strptime(assignment.end_time, '%H:%M')
            hours = (end - start).total_seconds() / 3600

            staff_totals[staff_name]['hours'] += hours
            staff_totals[staff_name]['count'] += 1

        # Write summary
        row = 2
        for staff_name, totals in sorted(staff_totals.items()):
            ws[f'A{row}'] = staff_name
            ws[f'B{row}'] = f"{totals['hours']:.1f}h"
            ws[f'C{row}'] = totals['count']
            row += 1


def import_to_database(parsed_data: Dict[str, Any], db_session) -> Dict[str, Any]:
    """
    Import parsed Excel data into the database.

    Args:
        parsed_data: Dictionary from parse_schedule_excel()
        db_session: SQLAlchemy database session

    Returns:
        Dictionary with created entities and statistics
    """
    from datetime import datetime, timedelta
    import uuid

    stats = {
        'students_created': 0,
        'staff_created': 0,
        'classes_created': 0,
        'care_times_created': 0,
        'work_hours_created': 0,
    }

    # Create classes first (needed for student FK)
    class_map = {}
    for class_data in parsed_data['classes']:
        school_class = SchoolClass(
            id=uuid.uuid4(),
            name=class_data['name'],
            grade_group=class_data['grade_group'],
            academic_year="2025/2026",
            active=True,
        )
        db_session.add(school_class)
        class_map[class_data['name']] = school_class
        stats['classes_created'] += 1

    db_session.flush()  # Get IDs for classes

    # Create staff
    staff_map = {}
    for staff_data in parsed_data['staff']:
        # Find or create staff
        existing_staff = db_session.query(Staff).filter_by(
            personal_number=staff_data['personal_number']
        ).first()

        if existing_staff:
            staff = existing_staff
        else:
            staff = Staff(
                id=uuid.uuid4(),
                personal_number=staff_data['personal_number'],
                first_name=staff_data['first_name'],
                last_name=staff_data['last_name'],
                role=staff_data['role'],
                care_certifications=staff_data.get('care_certifications', []),
                schedule_type=ScheduleType.FIXED,
                employment_start=datetime.now(),
                active=True,
            )
            db_session.add(staff)
            stats['staff_created'] += 1

        staff_map[f"{staff_data['first_name']} {staff_data['last_name']}"] = staff

        # Create work hours
        for weekday, work_time in staff_data.get('work_hours', {}).items():
            work_hour = WorkHour(
                id=uuid.uuid4(),
                staff_id=staff.id,
                weekday=weekday,
                week_number=0,  # Apply to all weeks
                start_time=work_time['start'],
                end_time=work_time['end'],
            )
            db_session.add(work_hour)
            stats['work_hours_created'] += 1

    db_session.flush()

    # Create students
    student_map = {}
    for student_data in parsed_data['students']:
        # Determine class assignment (simplified - assign based on grade_group)
        grade_group = student_data['grade_group']
        class_suffix = 'A'  # Default to first class in group
        class_name = f"Klass {'1-3' if grade_group == GradeGroup.GRADES_1_3 else '4-6'}{class_suffix}"

        class_id = class_map.get(class_name).id if class_name in class_map else None

        # Find or create student
        existing_student = db_session.query(Student).filter_by(
            personal_number=student_data['personal_number']
        ).first()

        if existing_student:
            student = existing_student
        else:
            student = Student(
                id=uuid.uuid4(),
                personal_number=student_data['personal_number'],
                first_name=student_data['first_name'],
                last_name=student_data['last_name'],
                class_id=class_id,
                grade=student_data['grade'],
                has_care_needs=student_data.get('has_care_needs', False),
                care_requirements=student_data.get('care_requirements', []),
                preferred_staff=[],  # Will populate below
                requires_double_staffing=student_data.get('requires_double_staffing', False),
                active=True,
            )
            db_session.add(student)
            stats['students_created'] += 1

        student_map[f"{student_data['first_name']} {student_data['last_name']}"] = student

        # Create care times
        for weekday, care_time in student_data.get('care_times', {}).items():
            # Set validity period (current week)
            today = datetime.now().date()
            week_start = today - timedelta(days=today.weekday())

            care_time_obj = CareTime(
                id=uuid.uuid4(),
                student_id=student.id,
                weekday=weekday,
                start_time=care_time['start'],
                end_time=care_time['end'],
                valid_from=datetime.combine(week_start, datetime.min.time()),
                valid_to=None,  # Indefinite
            )
            db_session.add(care_time_obj)
            stats['care_times_created'] += 1

    db_session.commit()

    return {
        'stats': stats,
        'students': list(student_map.values()),
        'staff': list(staff_map.values()),
        'classes': list(class_map.values()),
    }


class ExcelTemplateService:
    """
    Service for generating clean Excel templates for data import.

    Creates a structured template with separate sheets for:
    - Students
    - Staff
    - Classes
    - Care Times
    - Work Hours
    - Instructions (Swedish)
    """

    @staticmethod
    def generate_template() -> bytes:
        """
        Generate a clean Excel template for bulk data import.

        Returns:
            bytes: Excel file content
        """
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        from io import BytesIO

        wb = Workbook()

        # Remove default sheet
        if 'Sheet' in wb.sheetnames:
            del wb['Sheet']

        # Sheet 1: Instructions
        instructions_sheet = wb.create_sheet("📖 Instruktioner", 0)
        ExcelTemplateService._create_instructions_sheet(instructions_sheet)

        # Sheet 2: Students
        students_sheet = wb.create_sheet("👶 Elever", 1)
        ExcelTemplateService._create_students_sheet(students_sheet)

        # Sheet 3: Staff
        staff_sheet = wb.create_sheet("👤 Personal", 2)
        ExcelTemplateService._create_staff_sheet(staff_sheet)

        # Sheet 4: Classes
        classes_sheet = wb.create_sheet("📚 Klasser", 3)
        ExcelTemplateService._create_classes_sheet(classes_sheet)

        # Sheet 5: Care Times
        care_times_sheet = wb.create_sheet("⏰ Omsorgstider", 4)
        ExcelTemplateService._create_care_times_sheet(care_times_sheet)

        # Sheet 6: Work Hours
        work_hours_sheet = wb.create_sheet("💼 Arbetstider", 5)
        ExcelTemplateService._create_work_hours_sheet(work_hours_sheet)

        # Save to bytes
        output = BytesIO()
        wb.save(output)
        output.seek(0)

        return output.getvalue()

    @staticmethod
    def _create_instructions_sheet(sheet):
        """Create instructions sheet with Swedish guidance."""
        from openpyxl.styles import Font, PatternFill, Alignment

        # Title
        sheet['A1'] = "Instruktioner för Excel-import"
        sheet['A1'].font = Font(bold=True, size=16)

        instructions = [
            "",
            "Välkommen! Detta är mallen för att importera elever, personal och scheman i bulk.",
            "",
            "📋 STEG FÖR STEG:",
            "1. Fyll i arket '📚 Klasser' först - ange alla klasser",
            "2. Fyll i arket '👶 Elever' - namn, årskurs, klass",
            "3. Fyll i arket '👤 Personal' - namn, roll, certifieringar",
            "4. Fyll i arket '⏰ Omsorgstider' - när varje elev behöver omsorg",
            "5. Fyll i arket '💼 Arbetstider' - när varje personal arbetar",
            "6. Spara filen och ladda upp den i systemet",
            "",
            "⚠️ VIKTIGT:",
            "• Ändra INTE kolumnrubrikerna - systemet förväntar sig exakt dessa namn",
            "• Namn måste vara unika (förnamn + efternamn)",
            "• Tider ska vara i format HH:MM (t.ex. 08:00, 16:30)",
            "• Veckodagar: 0 = Måndag, 1 = Tisdag, 2 = Onsdag, 3 = Torsdag, 4 = Fredag",
            "• Årskurser: 0-6 (F-6)",
            "",
            "💡 TIPS:",
            "• Börja med en liten testgrupp (5-10 elever) för att testa importen",
            "• Kopiera/klistra in data från befintliga Excel-filer",
            "• Spara ofta!",
            "",
            "📞 HJÄLP:",
            "Om något går fel, kontrollera att:",
            "1. Alla obligatoriska fält är ifyllda",
            "2. Namn är unika (samma person ska inte finnas två gånger)",
            "3. Klassnamn matchar mellan arken",
            "4. Tider är i rätt format (HH:MM)",
            "",
            "Lycka till! 🚀",
        ]

        for i, text in enumerate(instructions, start=2):
            sheet[f'A{i}'] = text
            if text.startswith(("📋", "⚠️", "💡", "📞")):
                sheet[f'A{i}'].font = Font(bold=True, size=12)

        sheet.column_dimensions['A'].width = 100

    @staticmethod
    def _create_students_sheet(sheet):
        """Create students sheet with headers and example data."""
        from openpyxl.styles import Font, PatternFill

        # Headers
        headers = [
            "Förnamn *",
            "Efternamn *",
            "Årskurs (0-6) *",
            "Klass *",
            "Har vårdbehov (JA/NEJ)",
            "Vårdbehov (kommaseparerat)",
            "Kräver dubbelbemanning (JA/NEJ)",
        ]

        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")

        for col, header in enumerate(headers, start=1):
            cell = sheet.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font

        # Example data
        sheet.append([
            "Anna",
            "Andersson",
            "3",
            "3A",
            "JA",
            "Epilepsi, Allergi",
            "NEJ",
        ])

        sheet.append([
            "Bengt",
            "Bengtsson",
            "3",
            "3A",
            "NEJ",
            "",
            "NEJ",
        ])

        # Set column widths
        sheet.column_dimensions['A'].width = 18
        sheet.column_dimensions['B'].width = 15
        sheet.column_dimensions['C'].width = 15
        sheet.column_dimensions['D'].width = 15
        sheet.column_dimensions['E'].width = 12
        sheet.column_dimensions['F'].width = 25
        sheet.column_dimensions['G'].width = 30
        sheet.column_dimensions['H'].width = 30

    @staticmethod
    def _create_staff_sheet(sheet):
        """Create staff sheet with headers and example data."""
        from openpyxl.styles import Font, PatternFill

        headers = [
            "Förnamn *",
            "Efternamn *",
            "Roll *",
            "Certifieringar (kommaseparerat)",
            "Schematyp *",
        ]

        header_fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")

        for col, header in enumerate(headers, start=1):
            cell = sheet.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font

        # Example data
        sheet.append([
            "Karin",
            "Karlsson",
            "ELEVASSISTENT",
            "Epilepsi, Diabetes",
            "FAST",
        ])

        sheet.append([
            "Lars",
            "Larsson",
            "PEDAGOG",
            "",
            "FAST",
        ])

        # Add comments for valid values
        sheet['C2'].comment = openpyxl.comments.Comment(
            "Giltiga roller:\n- ELEVASSISTENT\n- PEDAGOG\n- FRITIDSPEDAGOG",
            "System"
        )

        sheet['E2'].comment = openpyxl.comments.Comment(
            "Giltiga schematyper:\n- FAST\n- TVÅVECKORS",
            "System"
        )

        # Set column widths
        for col in ['A', 'B']:
            sheet.column_dimensions[col].width = 18
        sheet.column_dimensions['C'].width = 20
        sheet.column_dimensions['D'].width = 35
        sheet.column_dimensions['E'].width = 15

    @staticmethod
    def _create_classes_sheet(sheet):
        """Create classes sheet with headers and example data."""
        from openpyxl.styles import Font, PatternFill

        headers = [
            "Klassnamn *",
            "Årskursgrupp *",
            "Ansvarig lärare (namn)",
            "Läsår *",
        ]

        header_fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
        header_font = Font(bold=True, color="000000")

        for col, header in enumerate(headers, start=1):
            cell = sheet.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font

        # Example data
        sheet.append(["F", "FÖRSKOLEKLASS", "", "2025/2026"])
        sheet.append(["1A", "LÅGSTADIUM", "Lars Larsson", "2025/2026"])
        sheet.append(["2A", "LÅGSTADIUM", "", "2025/2026"])
        sheet.append(["3A", "MELLANSTADIUM", "", "2025/2026"])

        # Add comment
        sheet['B2'].comment = openpyxl.comments.Comment(
            "Giltiga årskursgrupper:\n- FÖRSKOLEKLASS\n- LÅGSTADIUM (1-3)\n- MELLANSTADIUM (4-6)",
            "System"
        )

        # Set column widths
        sheet.column_dimensions['A'].width = 15
        sheet.column_dimensions['B'].width = 20
        sheet.column_dimensions['C'].width = 30
        sheet.column_dimensions['D'].width = 15

    @staticmethod
    def _create_care_times_sheet(sheet):
        """Create care times sheet with headers and example data."""
        from openpyxl.styles import Font, PatternFill

        headers = [
            "Elev Namn (Förnamn Efternamn) *",
            "Veckodag (0-4) *",
            "Starttid (HH:MM) *",
            "Sluttid (HH:MM) *",
        ]

        header_fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
        header_font = Font(bold=True, color="000000")

        for col, header in enumerate(headers, start=1):
            cell = sheet.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font

        # Example data
        sheet.append(["Anna Andersson", "0", "07:00", "16:00"])
        sheet.append(["Anna Andersson", "1", "07:00", "16:00"])
        sheet.append(["Anna Andersson", "2", "07:00", "16:00"])
        sheet.append(["Erik Eriksson", "0", "08:00", "17:00"])

        # Add comment
        sheet['B2'].comment = openpyxl.comments.Comment(
            "Veckodagar:\n0 = Måndag\n1 = Tisdag\n2 = Onsdag\n3 = Torsdag\n4 = Fredag",
            "System"
        )

        # Set column widths
        for col in ['A', 'B', 'C', 'D']:
            sheet.column_dimensions[col].width = 20

    @staticmethod
    def _create_work_hours_sheet(sheet):
        """Create work hours sheet with headers and example data."""
        from openpyxl.styles import Font, PatternFill

        headers = [
            "Personal Namn (Förnamn Efternamn) *",
            "Veckodag (0-4) *",
            "Starttid (HH:MM) *",
            "Sluttid (HH:MM) *",
            "Lunch start (HH:MM)",
            "Lunch slut (HH:MM)",
        ]

        header_fill = PatternFill(start_color="C5E0B4", end_color="C5E0B4", fill_type="solid")
        header_font = Font(bold=True, color="000000")

        for col, header in enumerate(headers, start=1):
            cell = sheet.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font

        # Example data
        sheet.append(["Karin Karlsson", "0", "07:00", "16:00", "12:00", "12:30"])
        sheet.append(["Karin Karlsson", "1", "07:00", "16:00", "12:00", "12:30"])
        sheet.append(["Lars Larsson", "0", "08:00", "17:00", "11:30", "12:00"])

        # Add comment
        sheet['B2'].comment = openpyxl.comments.Comment(
            "Veckodagar:\n0 = Måndag\n1 = Tisdag\n2 = Onsdag\n3 = Torsdag\n4 = Fredag",
            "System"
        )

        # Set column widths
        for col in ['A', 'B', 'C', 'D', 'E', 'F']:
            sheet.column_dimensions[col].width = 22
