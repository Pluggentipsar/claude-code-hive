"""
Generate realistic mock data Excel file for Kålgårdens anpassade grundskola.

Based on the real schedule ("Schema att maila Joel.xlsx"):
- 9 classes: 1-3A, 1-3B, 1-3C, 1-3D, 1-3E (lågstadium), 4-6A, 4-6B, 4-6C, 4-6D (mellanstadium)
- 32 students in the schedule (with omsorgstider), ~14 more without fritids
- Open hours: 06:00-18:00
- School hours grades 1-3: 08:30-13:30
- School hours grades 4-6: Mon/Wed 08:30-14:30, Tue/Thu 08:30-15:00, Fri 08:30-13:30
- Staff: elevassistenter, fritidspedagoger, pedagoger
- Some students have vårdbehov (special care needs)
- KTS students (Eliah, Jeremiah) handled separately
- Elevassistenter have 2-week rotating schedules with rolling Fridays
"""

from openpyxl.styles import Font, PatternFill
from openpyxl import Workbook

# ============================================================================
# DATA DEFINITIONS - Based on real schedule
# ============================================================================

# 9 classes matching the real schedule
CLASSES = [
    ("1-3A", "LÅGSTADIUM"),
    ("1-3B", "LÅGSTADIUM"),
    ("1-3C", "LÅGSTADIUM"),
    ("1-3D", "LÅGSTADIUM"),
    ("1-3E", "LÅGSTADIUM"),
    ("4-6A", "MELLANSTADIUM"),
    ("4-6B", "MELLANSTADIUM"),
    ("4-6C", "MELLANSTADIUM"),
    ("4-6D", "MELLANSTADIUM"),
]

# Students from the real schedule
# (first_name, last_name, grade, class, has_fritids, has_vårdbehov, vårdbehov_list, dubbelbemanning)
# Care times from Monday's sheet Start/Slut columns
STUDENTS = [
    # ---- LÅGSTADIUM ----

    # 1-3A: William, Niklas (vårdbehov), Gideon, Jeremia
    ("William", "Andersson", 1, "1-3A", True, False, "", False),
    ("Niklas", "Eriksson", 2, "1-3A", True, True, "Epilepsi, Medicinering", True),
    ("Gideon", "Johansson", 1, "1-3A", True, False, "", False),
    ("Jeremia", "Larsson", 3, "1-3A", True, False, "", False),

    # 1-3B: Albin, Loke, Dave, Ivan, Johan
    ("Albin", "Svensson", 2, "1-3B", True, False, "", False),
    ("Loke", "Pettersson", 1, "1-3B", True, False, "", False),
    ("Dave", "Nilsson", 3, "1-3B", True, False, "", False),
    ("Ivan", "Karlsson", 2, "1-3B", True, False, "", False),
    ("Johan", "Gustafsson", 3, "1-3B", True, False, "", False),

    # 1-3C: Ajan, Christelle, Sanarya
    ("Ajan", "Holm", 1, "1-3C", True, False, "", False),
    ("Christelle", "Fransson", 2, "1-3C", True, False, "", False),
    ("Sanarya", "Wallin", 3, "1-3C", True, False, "", False),

    # 1-3D: Aaryan, Muhammed
    ("Aaryan", "Lundqvist", 1, "1-3D", True, False, "", False),
    ("Muhammed", "Bergström", 2, "1-3D", True, False, "", False),

    # 1-3E: Kiebron, Elliot, Jonathan (låg), Nami
    ("Kiebron", "Nordström", 2, "1-3E", True, False, "", False),
    ("Elliot", "Lindgren", 1, "1-3E", True, False, "", False),
    ("Jonathan", "Ekström", 3, "1-3E", True, False, "", False),
    ("Nami", "Sandberg", 1, "1-3E", True, False, "", False),

    # Extra lågstadium students without fritids
    ("Elias", "Dahl", 1, "1-3A", False, False, "", False),
    ("Felicia", "Lindqvist", 2, "1-3B", False, False, "", False),
    ("Noah", "Björk", 3, "1-3C", False, False, "", False),
    ("Tilde", "Magnusson", 1, "1-3D", False, False, "", False),
    ("Adam", "Blom", 2, "1-3E", False, False, "", False),

    # ---- MELLANSTADIUM ----

    # 4-6A: Jonathan (mellan), Liam, Thomas, Olivia
    ("Jonathan M", "Borg", 4, "4-6A", True, False, "", False),
    ("Liam", "Sjöberg", 5, "4-6A", True, False, "", False),
    ("Thomas", "Henriksson", 6, "4-6A", True, False, "", False),
    ("Olivia", "Lund", 4, "4-6A", True, False, "", False),

    # 4-6B: Elsa, Abubaker
    ("Elsa", "Norberg", 5, "4-6B", True, False, "", False),
    ("Abubaker", "Åberg", 4, "4-6B", True, False, "", False),

    # 4-6C: Wiggo, Lowe, Tristan, Mari-Bell
    ("Wiggo", "Forsberg", 5, "4-6C", True, False, "", False),
    ("Lowe", "Engström", 4, "4-6C", True, False, "", False),
    ("Tristan", "Ström", 6, "4-6C", True, False, "", False),
    ("Mari-Bell", "Persson", 5, "4-6C", True, False, "", False),

    # 4-6D: Asta, Thilde
    ("Asta", "Holmberg", 4, "4-6D", True, False, "", False),
    ("Thilde", "Ahlström", 6, "4-6D", True, False, "", False),

    # Extra mellanstadium students without fritids
    ("Saga", "Hedlund", 4, "4-6A", False, False, "", False),
    ("Leon", "Sundberg", 5, "4-6B", False, False, "", False),
    ("Minna", "Wikström", 6, "4-6C", False, False, "", False),
    ("Filip", "Söderström", 4, "4-6D", False, False, "", False),

    # ---- KTS (Klass för träning och stimulans) ----
    # These students don't have class assignments in the regular sense
    ("Eliah", "Roos", 3, "1-3C", True, True, "Autism, Kommunikation", True),
    ("Jeremiah", "Sundén", 4, "4-6A", True, True, "Autism, Rörelsehinder", True),
]

# Staff from the real schedule
# Elevassistenter - matched to classes based on the Excel
ELEVASSISTENTER = [
    # --- Lågstadium staff ---
    # 1-3A primary staff
    ("Fadi A", "Ahmadi", "1-3A", "Epilepsi, Medicinering"),
    ("Jeanette", "Nordqvist", "1-3A", "Epilepsi"),
    ("Cia", "Samuelsson", "1-3A", ""),
    ("Emma", "Åkesson", "1-3A", "Medicinering"),

    # 1-3B primary staff
    ("Kim", "Häggström", "1-3B", ""),
    ("Kristina", "Berglind", "1-3B", ""),
    ("Maha", "Söderlund", "1-3B", ""),
    ("Stefan", "Jonasson", "1-3B", ""),
    ("Georgette", "Öhman", "1-3B", ""),

    # 1-3E primary staff
    ("Arafo", "Granberg", "1-3E", ""),
    ("Erica", "Isaksson", "1-3E", ""),
    ("Fadi K", "Khalil", "1-3E", ""),
    ("Halla", "Vestlund", "1-3E", ""),

    # 1-3C primary staff
    ("Helene", "Ljungberg", "1-3C", ""),
    ("Emelie K", "Karlsson", "1-3C", ""),
    ("Annelie", "Edlund", "1-3C", ""),

    # 1-3D primary staff
    ("Lena N", "Nyberg", "1-3D", "Medicinering"),
    ("Matilda", "Engberg", "1-3D", ""),

    # Lågstadium resource staff
    ("Jean", "Viklund", "1-3E", ""),
    ("Erik V", "Vikström", "1-3E", ""),
    ("Lotta", "Bergqvist", "1-3B", "Epilepsi"),
    ("Ida", "Hellberg", "1-3B", ""),

    # --- Mellanstadium staff ---
    # 4-6A primary staff
    ("Tobias", "Brandt", "4-6A", ""),
    ("Rozalija", "Sundqvist", "4-6A", ""),
    ("Carl H", "Holmgren", "4-6A", ""),

    # 4-6B primary staff
    ("Therese G", "Gustafsson", "4-6B", ""),
    ("Therese J", "Johansson", "4-6B", ""),

    # 4-6C primary staff
    ("Carl S", "Sandström", "4-6C", ""),
    ("Ruzmare", "Löfgren", "4-6C", ""),
    ("Magdalena", "Lindahl", "4-6C", ""),
    ("Ludwig", "Wahlström", "4-6C", ""),
    ("Stina", "Sjölund", "4-6C", ""),

    # 4-6D primary staff
    ("Angelica", "Eklund", "4-6D", ""),
    ("Susanne", "Nilsen", "4-6D", ""),
    ("Nawal", "Moberg", "4-6D", ""),
    ("Nimo", "Hägglund", "4-6D", ""),

    # Mellanstadium resource staff
    ("Eric", "Malmström", "4-6C", ""),
    ("Darin", "Strömberg", "4-6C", ""),
]

# Fritidspedagoger (from the real schedule)
FRITIDSPEDAGOGER = [
    ("Anna", "Lindström", "4-6D"),
    ("Anki F", "Falk", "1-3C"),
    ("Anna-Maria", "Mårtensson", "1-3B"),
    ("Gabriella", "Lundberg", None),
    ("Mange", "Carlsson", None),  # Pool
]

# Pedagoger (classroom teachers, not in the schedule but needed)
PEDAGOGER = [
    ("Maria", "Bergman", "1-3A"),
    ("Per", "Nordin", "1-3B"),
    ("Karin", "Åström", "1-3C"),
    ("Anders", "Öberg", "1-3D"),
    ("Birgitta", "Hellström", "1-3E"),
    ("Lars", "Dahlström", "4-6A"),
    ("Camilla", "Eliasson", "4-6B"),
    ("Henrik", "Lindblom", "4-6C"),
    ("Sandra", "Paulsson", "4-6D"),
]

# ============================================================================
# CARE TIMES - Omsorgstider from the real schedule
# ============================================================================

# Actual care times extracted from the Excel (Monday data as baseline)
# Format: (student_first_name, weekday_times_dict)
# Each day: (start, end)
CARE_TIMES_RAW = {
    # 1-3A
    "William": ("07:30", "16:30"),
    "Niklas": ("08:30", "15:30"),
    "Gideon": ("06:30", "16:40"),
    "Jeremia": ("06:00", "17:00"),
    # 1-3B
    "Albin": ("06:30", "13:30"),
    "Loke": ("08:00", "13:30"),
    "Dave": ("08:00", "17:00"),
    "Ivan": ("07:45", "17:00"),
    "Johan": ("07:45", "16:30"),
    # 1-3E
    "Kiebron": ("08:00", "16:30"),
    "Elliot": ("06:45", "17:10"),
    "Jonathan": ("07:30", "16:30"),
    "Nami": ("06:00", "16:00"),
    # 1-3C
    "Ajan": ("08:00", "13:30"),
    "Christelle": ("07:30", "16:00"),
    "Sanarya": ("08:00", "17:00"),
    # 1-3D
    "Aaryan": ("08:00", "16:30"),
    "Muhammed": ("08:30", "16:00"),
    # 4-6A
    "Jonathan M": ("07:45", "14:00"),
    "Liam": ("08:00", "14:45"),
    "Thomas": ("08:00", "16:30"),
    "Olivia": ("08:00", "15:10"),
    # 4-6B
    "Elsa": ("07:15", "14:30"),
    "Abubaker": ("08:00", "14:30"),
    # 4-6C
    "Wiggo": ("08:00", "16:00"),
    "Lowe": ("08:30", "15:20"),
    "Tristan": ("08:00", "16:30"),
    "Mari-Bell": ("08:00", "16:30"),
    # 4-6D
    "Asta": ("08:00", "16:00"),
    "Thilde": ("07:15", "16:15"),
    # KTS
    "Eliah": ("08:00", "16:00"),
    "Jeremiah": ("08:00", "16:00"),
}


def generate_care_times():
    """Generate omsorgstider based on actual schedule data."""
    care_times = []

    for student in STUDENTS:
        first_name = student[0]
        has_fritids = student[4]

        if not has_fritids:
            continue

        full_name = f"{first_name} {student[1]}"

        if first_name in CARE_TIMES_RAW:
            start, end = CARE_TIMES_RAW[first_name]
        else:
            # Default for students not in the schedule
            start, end = "07:30", "16:00"

        # Slight variation per day (real schedule varies slightly)
        for weekday in range(5):
            # Friday tends to end earlier
            if weekday == 4:
                end_hour = int(end.split(":")[0])
                end_min = end.split(":")[1]
                adj_hour = max(13, end_hour - 1)
                day_end = f"{adj_hour:02d}:{end_min}"
            else:
                day_end = end

            care_times.append((full_name, weekday, start, day_end))

    return care_times


# ============================================================================
# WORK HOURS - Staff work hours based on real schedule
# ============================================================================

# Staff work hours from the Excel (column L-P on Monday)
STAFF_HOURS_RAW = {
    # Lågstadium
    "Fadi A": ("05:50", "13:45"),
    "Jeanette": ("07:30", "16:15"),
    "Cia": ("08:00", "16:30"),
    "Emma": ("08:30", "17:00"),
    "Maha": ("08:00", "17:00"),
    "Stefan": ("08:00", "16:45"),
    "Kim": ("06:30", "13:45"),
    "Kristina": ("07:45", "16:15"),
    "Georgette": ("08:00", "17:00"),
    "Ida": ("06:30", "13:45"),
    "Arafo": ("07:30", "13:30"),
    "Erica": ("09:00", "16:30"),
    "Fadi K": ("08:15", "17:30"),
    "Halla": ("08:30", "14:30"),
    "Helene": ("07:30", "16:00"),
    "Emelie K": ("09:00", "17:00"),
    "Annelie": ("08:00", "16:00"),
    "Lena N": ("08:00", "16:30"),
    "Matilda": ("06:00", "13:30"),
    "Jean": ("08:00", "16:30"),
    "Erik V": ("06:45", "13:45"),
    "Lotta": ("07:45", "16:15"),
    # Mellanstadium
    "Tobias": ("08:00", "15:45"),
    "Rozalija": ("08:00", "15:45"),
    "Carl H": ("07:00", "15:45"),
    "Therese G": ("08:00", "15:45"),
    "Therese J": ("06:30", "15:45"),
    "Carl S": ("08:00", "16:30"),
    "Ruzmare": ("08:30", "17:00"),
    "Magdalena": ("08:30", "16:30"),
    "Ludwig": ("08:00", "16:00"),
    "Stina": ("08:30", "16:30"),
    "Angelica": ("07:00", "16:00"),
    "Susanne": ("09:00", "16:00"),
    "Nawal": ("08:00", "16:15"),
    "Nimo": ("07:15", "14:30"),
    "Eric": ("08:00", "16:00"),
    "Darin": ("08:30", "16:00"),
    # Fritidspedagoger
    "Anna": ("07:00", "16:00"),
    "Anki F": ("08:00", "17:00"),
    "Anna-Maria": ("08:00", "16:00"),
    "Gabriella": ("08:00", "16:30"),
    "Mange": ("08:00", "14:30"),
}


def generate_work_hours():
    """Generate work hours based on actual schedule data."""
    work_hours = []

    # Pedagoger: standard school hours
    for ped in PEDAGOGER:
        first, last, _ = ped
        full_name = f"{first} {last}"
        for weekday in range(5):
            work_hours.append((full_name, weekday, "07:45", "16:15", "11:30", "12:00"))

    # Elevassistenter: from the real schedule
    for ea in ELEVASSISTENTER:
        first, last, _, _ = ea
        full_name = f"{first} {last}"
        raw = STAFF_HOURS_RAW.get(first, ("08:00", "16:30"))
        start, end = raw

        for weekday in range(5):
            # Friday often ends earlier
            if weekday == 4:
                end_hour = int(end.split(":")[0])
                end_min = end.split(":")[1]
                fri_hour = max(13, end_hour - 1)
                fri_end = f"{fri_hour:02d}:{end_min}"
            else:
                fri_end = end

            # Lunch around midday
            start_hour = int(start.split(":")[0])
            lunch_start_h = max(11, min(13, start_hour + 4))
            lunch_start = f"{lunch_start_h:02d}:00"
            lunch_end = f"{lunch_start_h:02d}:30"

            work_hours.append((full_name, weekday, start, fri_end, lunch_start, lunch_end))

    # Fritidspedagoger
    for fp in FRITIDSPEDAGOGER:
        first, last, _ = fp
        full_name = f"{first} {last}"
        raw = STAFF_HOURS_RAW.get(first, ("08:00", "16:30"))
        start, end = raw

        for weekday in range(5):
            start_hour = int(start.split(":")[0])
            lunch_start_h = max(11, min(13, start_hour + 4))
            lunch_start = f"{lunch_start_h:02d}:00"
            lunch_end = f"{lunch_start_h:02d}:30"

            work_hours.append((full_name, weekday, start, end, lunch_start, lunch_end))

    return work_hours


# ============================================================================
# GENERATE EXCEL FILE
# ============================================================================

def create_excel():
    wb = Workbook()
    if 'Sheet' in wb.sheetnames:
        del wb['Sheet']

    header_blue = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_green = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
    header_yellow = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
    header_gray = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
    header_ltgreen = PatternFill(start_color="C5E0B4", end_color="C5E0B4", fill_type="solid")
    white_font = Font(bold=True, color="FFFFFF")
    black_font = Font(bold=True, color="000000")

    def write_headers(sheet, headers, fill, font):
        for col, header in enumerate(headers, start=1):
            cell = sheet.cell(row=1, column=col, value=header)
            cell.fill = fill
            cell.font = font

    # ---- Sheet 1: Klasser ----
    ws_classes = wb.create_sheet("Klasser", 0)
    write_headers(ws_classes,
        ["Klassnamn *", "Årskursgrupp *", "Ansvarig lärare (namn)", "Läsår *"],
        header_yellow, black_font)

    for klass_name, grade_group in CLASSES:
        teacher = ""
        for ped in PEDAGOGER:
            if ped[2] == klass_name:
                teacher = f"{ped[0]} {ped[1]}"
                break
        ws_classes.append([klass_name, grade_group, teacher, "2025/2026"])

    ws_classes.column_dimensions['A'].width = 15
    ws_classes.column_dimensions['B'].width = 20
    ws_classes.column_dimensions['C'].width = 30
    ws_classes.column_dimensions['D'].width = 15

    # ---- Sheet 2: Elever ----
    ws_students = wb.create_sheet("Elever", 1)
    write_headers(ws_students,
        ["Förnamn *", "Efternamn *", "Årskurs (0-6) *", "Klass *",
         "Har vårdbehov (JA/NEJ)", "Vårdbehov (kommaseparerat)",
         "Kräver dubbelbemanning (JA/NEJ)"],
        header_blue, white_font)

    for s in STUDENTS:
        first, last, grade, klass, has_fritids, has_vard, vard_list, dubbel = s
        ws_students.append([
            first, last, grade, klass,
            "JA" if has_vard else "NEJ",
            vard_list,
            "JA" if dubbel else "NEJ",
        ])

    for col, w in [('A', 15), ('B', 15), ('C', 18), ('D', 10),
                   ('E', 25), ('F', 30), ('G', 32)]:
        ws_students.column_dimensions[col].width = w

    # ---- Sheet 3: Personal ----
    ws_staff = wb.create_sheet("Personal", 2)
    write_headers(ws_staff,
        ["Förnamn *", "Efternamn *", "Roll *",
         "Certifieringar (kommaseparerat)", "Schematyp *"],
        header_green, white_font)

    for ped in PEDAGOGER:
        ws_staff.append([ped[0], ped[1], "PEDAGOG", "", "FAST"])

    for fp in FRITIDSPEDAGOGER:
        ws_staff.append([fp[0], fp[1], "FRITIDSPEDAGOG", "", "FAST"])

    for ea in ELEVASSISTENTER:
        first, last, _, certs = ea
        ws_staff.append([first, last, "ELEVASSISTENT", certs, "TVÅVECKORS"])

    for col, w in [('A', 18), ('B', 18), ('C', 20), ('D', 35), ('E', 15)]:
        ws_staff.column_dimensions[col].width = w

    # ---- Sheet 4: Omsorgstider ----
    ws_care = wb.create_sheet("Omsorgstider", 3)
    write_headers(ws_care,
        ["Elev Namn (Förnamn Efternamn) *", "Veckodag (0-4) *",
         "Starttid (HH:MM) *", "Sluttid (HH:MM) *"],
        header_gray, black_font)

    care_times = generate_care_times()
    for name, weekday, start, end in care_times:
        ws_care.append([name, weekday, start, end])

    for col in ['A', 'B', 'C', 'D']:
        ws_care.column_dimensions[col].width = 30

    # ---- Sheet 5: Arbetstider ----
    ws_work = wb.create_sheet("Arbetstider", 4)
    write_headers(ws_work,
        ["Personal Namn (Förnamn Efternamn) *", "Veckodag (0-4) *",
         "Starttid (HH:MM) *", "Sluttid (HH:MM) *",
         "Lunch start (HH:MM)", "Lunch slut (HH:MM)"],
        header_ltgreen, black_font)

    work_hours = generate_work_hours()
    for entry in work_hours:
        ws_work.append(list(entry))

    for col in ['A', 'B', 'C', 'D', 'E', 'F']:
        ws_work.column_dimensions[col].width = 30

    # Save
    output_path = "kalgarden_mockdata.xlsx"
    wb.save(output_path)

    # Print summary
    total_students = len(STUDENTS)
    fritids_students = sum(1 for s in STUDENTS if s[4])
    vard_students = sum(1 for s in STUDENTS if s[5])
    dubbel_students = sum(1 for s in STUDENTS if s[7])
    total_staff = len(PEDAGOGER) + len(FRITIDSPEDAGOGER) + len(ELEVASSISTENTER)

    print(f"Excel-fil skapad: {output_path}")
    print(f"  Klasser: {len(CLASSES)}")
    print(f"  Elever: {total_students} (varav {fritids_students} med fritids)")
    print(f"  Personal: {len(PEDAGOGER)} pedagoger + {len(FRITIDSPEDAGOGER)} fritidspedagoger + {len(ELEVASSISTENTER)} elevassistenter = {total_staff}")
    print(f"  Elever med vårdbehov: {vard_students}")
    print(f"  Elever med dubbelbemanning: {dubbel_students}")
    print(f"  Omsorgstider (rader): {len(care_times)}")
    print(f"  Arbetstider (rader): {len(work_hours)}")


if __name__ == "__main__":
    create_excel()
