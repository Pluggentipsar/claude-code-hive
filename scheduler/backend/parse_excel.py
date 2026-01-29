"""
Parse the Excel schedule file and extract students, staff, and schedule data.
"""

import openpyxl
from datetime import datetime, time
from typing import Dict, List, Set, Tuple
import re


def is_valid_name(name: str) -> bool:
    """Check if a string looks like a valid person name."""
    if not name or name in ['X', 'None', 'Skola', 'Fritids', 'Personal']:
        return False

    # Skip time strings
    if re.match(r'^\d{1,2}:\d{2}', name):
        return False

    # Skip obvious headers/notes
    if any(keyword in name for keyword in ['V.', 'mellan', 'arbetstid', 'KTS', 'eller', 'Fre1:']):
        return False

    # Skip if too long (likely a note)
    if len(name) > 30:
        return False

    return True


def parse_time(time_val) -> str:
    """Convert Excel time to HH:MM:SS string."""
    if isinstance(time_val, datetime):
        return time_val.strftime('%H:%M:%S')
    elif isinstance(time_val, time):
        return time_val.strftime('%H:%M:%S')
    elif isinstance(time_val, str):
        # Already a string, clean it up
        return time_val.strip()
    return str(time_val)


def parse_schedule_excel(file_path: str) -> Dict:
    """
    Parse the Excel schedule file.

    Returns:
        {
            'students': [{'name': str, 'class': str, 'care_times': [...]}],
            'staff': [{'name': str, 'work_hours': [...]}],
            'assignments': [...]
        }
    """
    wb = openpyxl.load_workbook(file_path)

    students_data = {}  # name -> student info
    staff_data = {}     # name -> staff info

    # Map weekday names to numbers (0=Monday)
    weekday_map = {
        'mån': 0,
        'tis': 1,
        'ons': 2,
        'tors': 3,
        'fre': 4,
    }

    # Process each sheet (one per weekday)
    for sheet_name in wb.sheetnames:
        # Extract weekday from sheet name (e.g., "mån 12 jan" -> "mån")
        weekday_str = sheet_name.split()[0].lower()
        weekday = weekday_map.get(weekday_str)

        if weekday is None:
            print(f"Warning: Could not parse weekday from sheet '{sheet_name}'")
            continue

        sheet = wb[sheet_name]
        print(f"\nProcessing {sheet_name} (weekday {weekday})...")

        current_class = None

        # Start from row 3 (row 1-2 are headers)
        for row_idx in range(3, sheet.max_row + 1):
            # Column 1: Class (if present)
            class_cell = sheet.cell(row=row_idx, column=1).value
            if class_cell and str(class_cell).strip():
                class_str = str(class_cell).strip()
                # Valid class format: digits-digits+letter (e.g., "1-3A", "4-6B")
                if re.match(r'^\d-\d[A-Z]$', class_str):
                    current_class = class_str

            # Column 2: Student name
            student_name = sheet.cell(row=row_idx, column=2).value
            if not student_name:
                continue

            student_name = str(student_name).strip()

            if not is_valid_name(student_name):
                continue

            # Column 4: Start time
            start_time = sheet.cell(row=row_idx, column=4).value
            # Column 5: End time
            end_time = sheet.cell(row=row_idx, column=5).value

            # Skip if no times or invalid times (X, None, etc.)
            if not start_time or not end_time:
                continue

            # Convert to string and check for invalid values
            start_str = str(start_time).strip()
            end_str = str(end_time).strip()

            if start_str in ['X', 'None', ''] or end_str in ['X', 'None', '']:
                continue

            # Initialize student if first time seeing them
            if student_name not in students_data:
                students_data[student_name] = {
                    'name': student_name,
                    'class': current_class,
                    'care_times': []
                }

            # Add care time for this weekday
            students_data[student_name]['care_times'].append({
                'weekday': weekday,
                'start_time': parse_time(start_str),
                'end_time': parse_time(end_str)
            })

            # Column 6: Fritids staff (before school)
            fritids_before = sheet.cell(row=row_idx, column=6).value
            if fritids_before:
                staff_name = str(fritids_before).strip().split('/')[0].strip()
                if is_valid_name(staff_name):
                    if staff_name not in staff_data:
                        staff_data[staff_name] = {
                            'name': staff_name,
                            'work_hours': [],
                            'assignments': []
                        }
                    staff_data[staff_name]['assignments'].append({
                        'student': student_name,
                        'weekday': weekday,
                        'period': 'fritids_before'
                    })

            # Column 9: Fritids staff (after school)
            fritids_after = sheet.cell(row=row_idx, column=9).value
            if fritids_after:
                staff_name = str(fritids_after).strip().split('/')[0].strip()
                if is_valid_name(staff_name):
                    if staff_name not in staff_data:
                        staff_data[staff_name] = {
                            'name': staff_name,
                            'work_hours': [],
                            'assignments': []
                        }
                    staff_data[staff_name]['assignments'].append({
                        'student': student_name,
                        'weekday': weekday,
                        'period': 'fritids_after'
                    })

            # Column 12: Staff name (detailed schedule section)
            detailed_staff = sheet.cell(row=row_idx, column=12).value
            if detailed_staff:
                staff_name = str(detailed_staff).strip().split('/')[0].strip()
                if is_valid_name(staff_name):
                    # Column 13: Staff start time
                    staff_start = sheet.cell(row=row_idx, column=13).value
                    # Column 14: Staff end time
                    staff_end = sheet.cell(row=row_idx, column=14).value
                    # Column 15: Lunch duration (e.g., 00:30:00)
                    lunch_duration = sheet.cell(row=row_idx, column=15).value

                    # Validate times
                    if staff_start and staff_end:
                        start_str = str(staff_start).strip()
                        end_str = str(staff_end).strip()

                        if start_str not in ['X', 'None', ''] and end_str not in ['X', 'None', '']:
                            if staff_name not in staff_data:
                                staff_data[staff_name] = {
                                    'name': staff_name,
                                    'work_hours': [],
                                    'assignments': []
                                }

                            # Add work hours for this weekday
                            staff_data[staff_name]['work_hours'].append({
                                'weekday': weekday,
                                'start_time': parse_time(start_str),
                                'end_time': parse_time(end_str),
                                'lunch_duration': parse_time(lunch_duration) if lunch_duration else '00:30:00'
                            })

    # Post-processing: Replicate Monday schedule to all weekdays for students
    # who don't have explicit schedules for other days
    print("\nPost-processing: Replicating Monday schedule to all weekdays...")

    monday_students = {}
    other_days_students = set()

    for name, student in students_data.items():
        for care_time in student['care_times']:
            if care_time['weekday'] == 0:  # Monday
                monday_students[name] = care_time
            else:
                other_days_students.add(name)

    # For students who appear on Monday but not on other days,
    # replicate their Monday schedule to Tue-Fri
    replicated_students = 0
    for name, monday_care in monday_students.items():
        if name not in other_days_students:
            # This student only appears on Monday - replicate to all weekdays
            student = students_data[name]
            # Add Tuesday-Friday with same times as Monday
            for weekday in range(1, 5):  # 1=Tue, 2=Wed, 3=Thu, 4=Fri
                student['care_times'].append({
                    'weekday': weekday,
                    'start_time': monday_care['start_time'],
                    'end_time': monday_care['end_time']
                })
            replicated_students += 1

    print(f"  Replicated Monday schedule to all weekdays for {replicated_students} students")
    print(f"  {len(other_days_students)} students have varying schedules across days")

    # Do the same for staff work hours
    monday_staff = {}
    other_days_staff = set()

    for name, staff in staff_data.items():
        for work_hour in staff.get('work_hours', []):
            if work_hour['weekday'] == 0:  # Monday
                monday_staff[name] = work_hour
            else:
                other_days_staff.add(name)

    replicated_staff = 0
    for name, monday_work in monday_staff.items():
        if name not in other_days_staff:
            staff = staff_data[name]
            # Add Tuesday-Friday with same times as Monday
            for weekday in range(1, 5):
                staff['work_hours'].append({
                    'weekday': weekday,
                    'start_time': monday_work['start_time'],
                    'end_time': monday_work['end_time'],
                    'lunch_duration': monday_work['lunch_duration']
                })
            replicated_staff += 1

    print(f"  Replicated Monday work hours to all weekdays for {replicated_staff} staff")
    print(f"  {len(other_days_staff)} staff have varying schedules across days")

    result = {
        'students': list(students_data.values()),
        'staff': list(staff_data.values())
    }

    print(f"\n\nPARSING COMPLETE:")
    print(f"  Students: {len(result['students'])}")
    print(f"  Staff: {len(result['staff'])}")

    return result


if __name__ == '__main__':
    data = parse_schedule_excel('../../Schema att maila Joel.xlsx')

    print("\n\nSTUDENTS:")
    print("=" * 80)
    for student in sorted(data['students'], key=lambda s: (s['class'] or '', s['name'])):
        care_days = len(student['care_times'])
        print(f"  {student['name']:20s} | Class: {student['class'] or 'N/A':5s} | Care: {care_days} days")

    print("\n\nSTAFF:")
    print("=" * 80)
    for staff in sorted(data['staff'], key=lambda s: s['name']):
        assignments = len(staff.get('assignments', []))
        print(f"  {staff['name']:20s} | Assignments: {assignments}")
